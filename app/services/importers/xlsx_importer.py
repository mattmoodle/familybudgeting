from pathlib import Path

from openpyxl import load_workbook

from app.services.importers.base import ImportedRow, StatementImporter
from app.services.importers.common import choose_columns, parse_amount, parse_date


class XlsxImporter(StatementImporter):
    def parse(self, path: Path) -> list[ImportedRow]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            return []
        headers = [str(v or "") for v in all_rows[0]]
        date_col, value_col, desc_col, amount_col, credit_col = choose_columns(headers)
        if date_col is None or desc_col is None or amount_col is None:
            raise ValueError("Unable to identify date/description/amount columns in XLSX")

        result: list[ImportedRow] = []
        for row in all_rows[1:]:
            if len(row) <= max(date_col, desc_col, amount_col):
                continue
            booked = parse_date(row[date_col])
            amount = parse_amount(row[amount_col])
            if credit_col is not None and credit_col < len(row):
                credit = parse_amount(row[credit_col])
                if credit not in (None, 0):
                    amount = abs(credit)
                elif amount is not None:
                    amount = -abs(amount)
            if booked is None or amount is None:
                continue
            value_on = parse_date(row[value_col]) if value_col is not None and value_col < len(row) else None
            result.append(ImportedRow(booked, str(row[desc_col] or "").strip(), amount, value_on, raw_data=str(row)))
        return result
